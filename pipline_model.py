import os
import toad
import warnings
import numpy as np
import pandas as pd
import scorecardpy as sc
import statsmodels.api as sm
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.ensemble import GradientBoostingClassifier
from toad.plot import bin_plot, proportion_plot, corr_plot, badrate_plot
from sklearn.base import BaseEstimator, TransformerMixin, ClassifierMixin
from statsmodels.stats.outliers_influence import variance_inflation_factor

from feature_bins import *


warnings.filterwarnings("ignore")
pd.set_option('display.width', 5000)
plt.rcParams["font.sans-serif"]=["SimHei"] #设置字体
plt.rcParams["axes.unicode_minus"]=False #该语句解决图像中的“-”负号的乱码问题


class FeatureSelection(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", empty=0.95, iv=0.02, corr=0.7, exclude=None, return_drop=True, identical=0.95, remove=None, engine="scorecardpy", target_rm=False):
        self.engine = engine
        self.target = target
        self.empty = empty
        self.identical = identical
        self.iv = iv
        self.corr = corr
        self.exclude = exclude
        self.remove = remove
        self.return_drop = return_drop
        self.target_rm = target_rm
        self.select_columns = None
        self.dropped = None
    
    def fit(self, x, y=None):
        if self.engine == "toad":
            selected = toad.selection.select(x, target=self.target, empty=self.empty, iv=self.iv, corr=self.corr, exclude=self.exclude, return_drop=self.return_drop)
        else:
            selected = sc.var_filter(x, y=self.target, iv_limit=self.iv, missing_limit=self.empty, identical_limit=self.identical, var_rm=self.remove, var_kp=self.exclude, return_rm_reason=self.return_drop)
            
        if self.return_drop and isinstance(selected, dict):
            self.dropped = selected["rm"]
            self.select_columns = list(selected["dt"].columns)
        elif self.return_drop and isinstance(selected, (tuple, list)):
            self.dropped = pd.DataFrame([(feature, reason) for reason, features in selected[1].items() for feature in features], columns=["variable", "rm_reason"])
            self.select_columns = list(selected[0].columns)
        else:
            self.select_columns = list(selected.columns)
        
        if self.target_rm and target in self.select_columns:
            self.select_columns.remove(target)
            
        return self
        
    def transform(self, x, y=None):
        # if self.engine == "toad":
        #     selected = toad.selection.select(x, target=self.target, empty=self.empty, iv=self.iv, corr=self.corr, exclude=self.exclude, return_drop=self.return_drop)
        # else:
        #     selected = sc.var_filter(x, y=self.target, iv_limit=self.iv, missing_limit=self.empty, identical_limit=self.identical, var_rm=self.remove, var_kp=self.exclude, return_rm_reason=self.return_drop)
            
        # if self.return_drop and isinstance(selected, dict):
        #     self.dropped = selected["rm"]
        #     return selected["dt"]
        # elif self.return_drop and isinstance(selected, (tuple, list)):
        #     self.dropped = pd.DataFrame([(feature, reason) for reason, features in selected[1].items() for feature in features], columns=["variable", "rm_reason"])
        #     return selected[0]
        # else:
        #     return selected
        return x[[col for col in self.select_columns if col in x.columns]]
    
    
class Combiner(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", method='chi', empty_separate=False, min_samples=0.05, n_bins=None, rules={}):
        self.combiner = toad.transform.Combiner()
        self.method = method
        self.empty_separate = empty_separate
        self.target = target
        self.min_samples = min_samples
        self.n_bins = n_bins
        self.rules = rules
    
    def fit(self, x, y=None):
        self.combiner.fit(x, y=self.target, method=self.method, min_samples=self.min_samples, n_bins=self.n_bins)
        self.update(self.rules)
        return self
    
    def transform(self, x, y=None, labels=False):
        return self.combiner.transform(x, labels=labels)
    
    def update(self, rules):
        if isinstance(rules, dict):
            self.combiner.update(self.rules)
            
    def bin_plot(self, data, x, rule=None, labels=True, result=False):
        if rule:
            if isinstance(rule, list):
                rule = {x: rule}
            self.combiner.update(rule)
            
        bin_plot(self.combiner.transform(data, labels=labels), x=x, target=self.target)
        
        if result:
            return self.combiner.export()[x]
        
    def proportion_plot(self, x, transform=False, labels=False):
        if transform:
            x = self.combiner.transform(x, labels=labels)
        proportion_plot(x)
        
    def corr_plot(self, data, transform=False, figure_size=(20, 15)):
        if transform:
            data = self.combiner.transform(data, labels=False)
        
        corr_plot(data, figure_size=figure_size)
        
    def badrate_plot(self, data, date_column, feature, labels=True):
        badrate_plot(self.combiner.transform(data[[date_column, feature, self.target]], labels=labels), target=self.target, x=date_column, by=feature)
        
        
class WOETransformer(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", exclude=None):
        self.target = target
        self.exclude = exclude if isinstance(exclude, list) else [exclude] if exclude else []
        self.transformer = toad.transform.WOETransformer()
        
    def fit(self, x, y=None):
        self.transformer.fit(x.drop(columns=self.exclude + [self.target]), x[self.target])
        return self

    def transform(self, x, y=None):
        return self.transformer.transform(x)
    
    
class StepwiseSelection(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", estimator="ols", direction="both", criterion="aic", max_iter=None, return_drop=True, exclude=None, intercept=True, p_value_enter=0.2, p_remove=0.01, p_enter=0.01, target_rm=False):
        self.target = target
        self.intercept = intercept
        self.p_value_enter = p_value_enter
        self.p_remove = p_remove
        self.p_enter = p_enter
        self.estimator = estimator
        self.direction = direction
        self.criterion = criterion
        self.max_iter = max_iter
        self.return_drop = return_drop
        self.target_rm = target_rm
        self.exclude = exclude
        self.select_columns = None
        self.dropped = None
    
    def fit(self, x, y=None):
        selected = toad.selection.stepwise(x, target=self.target, estimator=self.estimator, direction=self.direction, criterion=self.criterion, exclude=self.exclude, intercept=self.intercept, p_value_enter=self.p_value_enter, 
                                           p_remove=self.p_remove, p_enter=self.p_enter, return_drop=self.return_drop)
        if self.return_drop:
            self.dropped = pd.DataFrame([(col, "stepwise") for col in selected[1]], columns=["variable", "rm_reason"])
            selected = selected[0]
        
        self.select_columns = list(selected.columns)
        
        if self.target_rm and target in self.select_columns:
            self.select_columns.remove(target)
        
        return self
        
    def transform(self, x, y=None):
        return x[[col for col in self.select_columns if col in x.columns]]
    
    
class LogisticClassifier(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", intercept=True, ):
        self.intercept = intercept
        self.target = target
        self.classifier = None
        self.corr = None
        self.vif = None
        self.coef_normalization = None
        self.feature_names_ = None
        self.feature_importances_ = None
    
    def fit(self, x, y=None, vif=True, corr=True, normalization=True):
        self.feature_names_ = list(x.drop(columns=[self.target]).columns)
        self.feature_importances_ = self.feature_importances(x)
        
        if vif:
            self.vif = self.VIF(x)
            
        if normalization:
            _x = x.drop(columns=[self.target]).apply(lambda x: (x - np.mean(x)) / np.std(x))
            _y = x[self.target]
            lr_normalization = sm.Logit(_y, sm.add_constant(_x) if self.intercept else _x).fit()
            self.coef_normalization = pd.DataFrame(lr_normalization.params, columns=["coef_normalization"])
            
        if corr:
            self.corr = x.drop(columns=[self.target]).corr()
            
        if self.intercept:
            x = sm.add_constant(x)
        
        self.classes_ = x[self.target].unique()
        self.classifier = sm.Logit(x[self.target], x.drop(columns=[self.target])).fit()
        
        return self
    
    def transform(self, x):
        if self.intercept:
            x = sm.add_constant(x)
        
        return self.classifier.predict(x)
    
    def predict(self, x):
        return self.transform(x)
    
    def summary(self):
        describe = self.classifier.summary2()
        return describe
    
    def feature_importances(self, x):
        params = {
            "n_estimators": 256,
            "max_depth": 4,
            "min_samples_split": 5,
            "learning_rate": 1e-3,
            "loss": "deviance",
            "subsample": 0.9,
        }
        feature_importances_ = GradientBoostingClassifier(**params).fit(x.drop(columns=[self.target]), x[self.target]).feature_importances_
        return pd.DataFrame(feature_importances_, index=self.feature_names_, columns=["feature_importances"])
        
    def VIF(self, x):
        if self.intercept:
            x = sm.add_constant(x)
        
        x = x.drop(columns=[target])
        columns = x.columns
        vif = pd.DataFrame({"VIF": [variance_inflation_factor(np.matrix(x), i) for i in range(len(columns))]}, index=columns)
        
        return vif
    
    def WALD(self):
        return self.classifier.wald_test_terms().table[["statistic", "pvalue"]].rename(columns={"pvalue": "wald_test_pvalue", "statistic": "wald_test_statistic"})
    
    def report(self):
        return self.classifier.summary2().tables[1].join([self.coef_normalization, self.WALD(), self.vif, self.feature_importances_]), self.classifier.summary2().tables[0], self.corr
    
    def summary_save(self, excel_name="逻辑回归模型拟合效果.xlsx", sheet_name="逻辑回归拟合效果"):
        writer = pd.ExcelWriter(excel_name, engine='openpyxl')
        
        coef_report, summary_report, corr_report = self.report()
        summary_report.columns = ["逻辑回归模型拟合效果"] * summary_report.shape[1]
        summary_report.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startcol=0, startrow=2)
        coef_report.reset_index().rename(columns={"index": "variable"}).to_excel(writer, sheet_name=sheet_name, index=False, header=True, startcol=0, startrow=summary_report.shape[0] + 4)
        corr_report.to_excel(writer, sheet_name=sheet_name, index=True, header=True, startcol=0, startrow=summary_report.shape[0] + coef_report.shape[0] + 7)
        
        writer.save()
        writer.close()
        
        if os.path.exists(excel_name):
            workbook = load_workbook(excel_name)
            worksheet = workbook.get_sheet_by_name(sheet_name)
            worksheet["A1"].value = "逻辑回归模型报告"
            worksheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(f"A1:L1")
            
            workbook.save(excel_name)
            workbook.close()
        
        render_excel(excel_name, sheet_name=sheet_name, max_column_width=25, merge_rows=np.cumsum([1, len(summary_report), 2, len(coef_report) + 1, 2, len(corr_report) + 1]).tolist())
    
    
class ScoreCard(TransformerMixin, BaseEstimator, ClassifierMixin):
    
    def __init__(self, target="target", pdo=60, rate=2, base_odds=35, base_score=750, combiner={}, transer=None, penalty='l2', C=1.0, fit_intercept=True, class_weight="balanced", random_state=None, 
                 solver="lbfgs", max_iter=100, verbose=0, warm_start=False, n_jobs=None, l1_ratio=None, ):
        self.target = target
        self.combiner = combiner
        self.transer = transer
        self.base_score = base_score
        self.base_odds = base_odds
        self.rate = rate
        self.pdo = pdo
        self.penalty = penalty
        self.C = C
        self.fit_intercept = fit_intercept
        self.class_weight = class_weight
        self.random_state = random_state
        self.solver = solver
        self.max_iter = max_iter
        self.verbose = verbose
        self.warm_start = warm_start
        self.n_jobs = n_jobs
        self.l1_ratio = l1_ratio
        self.card = toad.ScoreCard(combiner=self.combiner, transer=self.transer, pdo=self.pdo, rate=self.rate, base_odds=self.base_odds, base_score=self.base_score, penalty=self.penalty, C=self.C, 
                                   fit_intercept=self.fit_intercept, class_weight=self.class_weight, random_state=self.random_state, solver=self.solver, max_iter=self.max_iter, verbose=self.verbose, 
                                   warm_start=self.warm_start, n_jobs=self.n_jobs, l1_ratio=self.l1_ratio, )
        
    def fit(self, x, y=None):
        self.classes_ = x[self.target].unique()
        self.card.fit(x.drop(columns=[self.target]), x[self.target])
        return self
    
    def transform(self, x):
        return self.card.predict(x)
    
    def predict(self, x):
        return self.card.predict(x)
    
    def export(self):
        return self.card.export()
    
    def KS_bucket(self, y_pred, y_true, bucket=10, method="quantile"):
        return toad.metrics.KS_bucket(y_pred, y_true, bucket=bucket, method=method)
    
    def KS(self, y_pred, y_true):
        return toad.metrics.KS(y_pred, y_true)
    
    def AUC(self, y_pred, y_true):
        return toad.metrics.AUC(y_pred, y_true)
    
    def perf_eva(self, y_pred, y_true, title="", plot_type=["ks", "roc"]):
        return sc.perf_eva(y_true, y_pred, title=title, plot_type=plot_type)
    
    def PSI(self, y_pred_train, y_pred_oot):
        return toad.metrics.PSI(y_pred_train, y_pred_oot)
    
    def perf_psi(self, y_pred_train, y_pred_oot, y_true_train, y_true_oot, keys=["train", "test"], x_limits=None, x_tick_break=50, show_plot=True, return_distr_dat=False):
        return sc.perf_psi(
            score = {keys[0]: y_pred_train, keys[1]: y_pred_oot},
            label = {keys[0]: y_true_train, keys[1]: y_true_oot},
            x_limits = x_limits,
            x_tick_break = x_tick_break,
            show_plot = show_plot,
            return_distr_dat = return_distr_dat,
        )
    

if __name__ == "__main__":
    target = "creditability"
    data = sc.germancredit()
    data[target] = data[target].map({"good": 0, "bad": 1})
    
    train, test = train_test_split(data, test_size=0.3, shuffle=True, stratify=data[target])
    
    # selection = FeatureSelection(target=target, engine="toad", return_drop=True, corr=0.9, iv=0.01)
    # train = selection.fit_transform(train)
    
    # combiner = Combiner(min_samples=0.2, empty_separate=True, target=target)
    # combiner.fit(train)
    # train = combiner.transform(train)
    
    # transformer = WOETransformer(target=target)
    # train = transformer.fit_transform(train)
    
    # stepwise = StepwiseSelection(target=target)
    # train = stepwise.fit_transform(train)
    
    feature_pipeline = Pipeline([
        ("preprocessing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("combiner", Combiner(target=target, min_samples=0.2)),
        ("transform", WOETransformer(target=target)),
        ("processing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("stepwise", StepwiseSelection(target=target, target_rm=False)),
        # ("logistic", LogisticClassifier()),
        # ("stepwise", StepwiseSelection(target=target, target_rm=True)),
        # ("logistic", LogisticRegression()),
    ])

    # params_grid = {
    #     "logistic__C": [i / 1. for i in range(1, 10, 2)],
    #     "logistic__penalty": ["l2"],
    #     "logistic__class_weight": [None, "balanced"], # + [{1: i / 10.0, 0: 1 - i / 10.0} for i in range(1, 10)],
    #     "logistic__max_iter": [100],
    #     "logistic__solver": ["sag"] # ["liblinear", "sag", "lbfgs", "newton-cg"],
    # }
    
    # clf = GridSearchCV(feature_pipeline, params_grid, cv=5, scoring='roc_auc', verbose=-1, n_jobs=2, return_train_score=True)
    # clf.fit(train, train[target])
    
    # y_pred_train = clf.best_estimator_.predict(train)
    # y_pred_test = clf.best_estimator_.predict(test)
    
    # print(clf.best_params_)
    # print("train: ", toad.metrics.KS(y_pred_train, train[target]), toad.metrics.AUC(y_pred_train, train[target]))
    # print("test: ", toad.metrics.KS(y_pred_test, test[target]), toad.metrics.AUC(y_pred_test, test[target]))
    
    woe_train = feature_pipeline.fit_transform(train)
    woe_test = feature_pipeline.transform(test)
    
    lr = LogisticClassifier(target=target)
    lr.fit(woe_train)
    lr.summary_save()

    cols = list(filter(lambda x: x != target, feature_pipeline.__getitem__(0).select_columns))
    combiner = feature_pipeline.__getitem__(1).combiner
    transformer = feature_pipeline.__getitem__(2).transformer
    
    feature_describe = pd.read_excel("变量字典及字段解释.xlsx", sheet_name="数据字段表", header=0, engine="openpyxl", usecols=[0, 1])
    feature_describe = feature_describe.drop_duplicates(subset=["变量名称"], keep="last")
    feature_dict = dict(zip(feature_describe["变量名称"], feature_describe["含义"]))

    # 保存结果至 EXCEL 文件
    output_excel_name = "指标有效性验证.xlsx"
    output_sheet_name = "指标有效性"
    tables = {}
    merge_row_number = []

    for feature in cols:
        table = feature_bin_stats(train, feature, feature_dict=feature_dict, rules={}, target=target)
        df_psi = cal_psi(train[[feature, target]], test[[feature, target]], feature, combiner=combiner)
        
        table = table.merge(df_psi, on="分箱", how="left")
        
        feature_bin = combiner.export()[feature]
        feature_bin_dict = format_bins(np.array(feature_bin))
        table["分箱"] = table["分箱"].map(feature_bin_dict)
        
        merge_row_number.append(len(table))
        tables[feature] = table

    merge_row_number = np.cumsum(merge_row_number).tolist()
    feature_table = pd.concat(tables, ignore_index=True).round(6)
    feature_table["分档WOE值"] = feature_table["分档WOE值"].fillna(np.inf)
    feature_table.to_excel(output_excel_name, sheet_name=output_sheet_name, index=False, header=True, startcol=0, startrow=0)

    render_excel(output_excel_name, sheet_name=output_sheet_name, conditional_columns=["J", "N"], freeze="D2", merge_rows=merge_row_number, percent_columns=[5, 7, 9, 10])
        
    score_card = ScoreCard(target=target, combiner=combiner, transer=transformer, )
    score_card.fit(woe_train)
    
    
    data["score"] = score_card.transform(data)
    
    print(score_card.KS_bucket(data["score"], data[target]))
    pt = score_card.perf_eva(data["score"], data[target], title="train")
    
    print(score_card.KS(data["score"], data[target]), score_card.AUC(data["score"], data[target]))
    