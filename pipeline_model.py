import os
import toad
import scipy
import warnings
import numpy as np
import pandas as pd
import scorecardpy as sc
import statsmodels.api as sm
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from IPython.display import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.utils.validation import check_is_fitted
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import GradientBoostingClassifier
from toad.plot import bin_plot, proportion_plot, corr_plot, badrate_plot
from sklearn.base import BaseEstimator, TransformerMixin, ClassifierMixin
from statsmodels.stats.outliers_influence import variance_inflation_factor

from feature_bins import *


warnings.filterwarnings("ignore")
pd.set_option('display.width', 5000)
plt.rcParams["font.sans-serif"]=["SimHei"] #设置字体
plt.rcParams["axes.unicode_minus"]=False #该语句解决图像中的“-”负号的乱码问题


class FeatureSelection(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", empty=0.95, iv=0.02, corr=0.7, exclude=None, return_drop=True, identical=0.95, remove=None, engine="scorecardpy", target_rm=False):
        self.engine = engine
        self.target = target
        self.empty = empty
        self.identical = identical
        self.iv = iv
        self.corr = corr
        self.exclude = exclude
        self.remove = remove
        self.return_drop = return_drop
        self.target_rm = target_rm
        self.select_columns = None
        self.dropped = None
    
    def fit(self, x, y=None):
        if self.engine == "toad":
            selected = toad.selection.select(x, target=self.target, empty=self.empty, iv=self.iv, corr=self.corr, exclude=self.exclude, return_drop=self.return_drop)
        else:
            selected = sc.var_filter(x, y=self.target, iv_limit=self.iv, missing_limit=self.empty, identical_limit=self.identical, var_rm=self.remove, var_kp=self.exclude, return_rm_reason=self.return_drop)
            
        if self.return_drop and isinstance(selected, dict):
            self.dropped = selected["rm"]
            self.select_columns = list(selected["dt"].columns)
        elif self.return_drop and isinstance(selected, (tuple, list)):
            self.dropped = pd.DataFrame([(feature, reason) for reason, features in selected[1].items() for feature in features], columns=["variable", "rm_reason"])
            self.select_columns = list(selected[0].columns)
        else:
            self.select_columns = list(selected.columns)
        
        if self.target_rm and target in self.select_columns:
            self.select_columns.remove(target)
            
        return self
        
    def transform(self, x, y=None):
        # if self.engine == "toad":
        #     selected = toad.selection.select(x, target=self.target, empty=self.empty, iv=self.iv, corr=self.corr, exclude=self.exclude, return_drop=self.return_drop)
        # else:
        #     selected = sc.var_filter(x, y=self.target, iv_limit=self.iv, missing_limit=self.empty, identical_limit=self.identical, var_rm=self.remove, var_kp=self.exclude, return_rm_reason=self.return_drop)
            
        # if self.return_drop and isinstance(selected, dict):
        #     self.dropped = selected["rm"]
        #     return selected["dt"]
        # elif self.return_drop and isinstance(selected, (tuple, list)):
        #     self.dropped = pd.DataFrame([(feature, reason) for reason, features in selected[1].items() for feature in features], columns=["variable", "rm_reason"])
        #     return selected[0]
        # else:
        #     return selected
        return x[[col for col in self.select_columns if col in x.columns]]
    
    
class Combiner(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", method='chi', empty_separate=False, min_samples=0.05, n_bins=None, rules={}):
        self.combiner = toad.transform.Combiner()
        self.method = method
        self.empty_separate = empty_separate
        self.target = target
        self.min_samples = min_samples
        self.n_bins = n_bins
        self.rules = rules
    
    def fit(self, x, y=None):
        self.combiner.fit(x, y=self.target, method=self.method, min_samples=self.min_samples, n_bins=self.n_bins)
        self.update(self.rules)
        return self
    
    def transform(self, x, y=None, labels=False):
        return self.combiner.transform(x, labels=labels)
    
    def update(self, rules):
        if isinstance(rules, dict):
            self.combiner.update(self.rules)
            
    def bin_plot(self, data, x, rule=None, labels=True, result=False):
        if rule:
            if isinstance(rule, list):
                rule = {x: rule}
            self.combiner.update(rule)
            
        bin_plot(self.combiner.transform(data, labels=labels), x=x, target=self.target)
        
        if result:
            return self.combiner.export()[x]
        
    def proportion_plot(self, x, transform=False, labels=False):
        if transform:
            x = self.combiner.transform(x, labels=labels)
        proportion_plot(x)
        
    def corr_plot(self, data, transform=False, figure_size=(20, 15)):
        if transform:
            data = self.combiner.transform(data, labels=False)
        
        corr_plot(data, figure_size=figure_size)
        
    def badrate_plot(self, data, date_column, feature, labels=True):
        badrate_plot(self.combiner.transform(data[[date_column, feature, self.target]], labels=labels), target=self.target, x=date_column, by=feature)
        
        
class WOETransformer(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", exclude=None):
        self.target = target
        self.exclude = exclude if isinstance(exclude, list) else [exclude] if exclude else []
        self.transformer = toad.transform.WOETransformer()
        
    def fit(self, x, y=None):
        self.transformer.fit(x.drop(columns=self.exclude + [self.target]), x[self.target])
        return self

    def transform(self, x, y=None):
        return self.transformer.transform(x)
    
    
class StepwiseSelection(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", estimator="ols", direction="both", criterion="aic", max_iter=None, return_drop=True, exclude=None, intercept=True, p_value_enter=0.2, p_remove=0.01, p_enter=0.01, target_rm=False):
        self.target = target
        self.intercept = intercept
        self.p_value_enter = p_value_enter
        self.p_remove = p_remove
        self.p_enter = p_enter
        self.estimator = estimator
        self.direction = direction
        self.criterion = criterion
        self.max_iter = max_iter
        self.return_drop = return_drop
        self.target_rm = target_rm
        self.exclude = exclude
        self.select_columns = None
        self.dropped = None
    
    def fit(self, x, y=None):
        selected = toad.selection.stepwise(x, target=self.target, estimator=self.estimator, direction=self.direction, criterion=self.criterion, exclude=self.exclude, intercept=self.intercept, p_value_enter=self.p_value_enter, 
                                           p_remove=self.p_remove, p_enter=self.p_enter, return_drop=self.return_drop)
        if self.return_drop:
            self.dropped = pd.DataFrame([(col, "stepwise") for col in selected[1]], columns=["variable", "rm_reason"])
            selected = selected[0]
        
        self.select_columns = list(selected.columns)
        
        if self.target_rm and target in self.select_columns:
            self.select_columns.remove(target)
        
        return self
        
    def transform(self, x, y=None):
        return x[[col for col in self.select_columns if col in x.columns]]
    
    
class LogisticClassifier(TransformerMixin, BaseEstimator):
    
    def __init__(self, target="target", intercept=True, ):
        self.intercept = intercept
        self.target = target
        self.classifier = None
        self.corr = None
        self.vif = None
        self.coef_normalization = None
        self.feature_names_ = None
        self.feature_importances_ = None
    
    def fit(self, x, y=None, vif=True, corr=True, normalization=True):
        self.feature_names_ = list(x.drop(columns=[self.target]).columns)
        self.feature_importances_ = self.feature_importances(x)
        
        if vif:
            self.vif = self.VIF(x)
            
        if normalization:
            _x = x.drop(columns=[self.target]).apply(lambda x: (x - np.mean(x)) / np.std(x))
            _y = x[self.target]
            lr_normalization = sm.Logit(_y, sm.add_constant(_x) if self.intercept else _x).fit()
            self.coef_normalization = pd.DataFrame(lr_normalization.params, columns=["coef_normalization"])
            
        if corr:
            self.corr = x.drop(columns=[self.target]).corr()
            
        if self.intercept:
            x = sm.add_constant(x)
        
        self.classes_ = x[self.target].unique()
        self.classifier = sm.Logit(x[self.target], x.drop(columns=[self.target])).fit()
        
        return self
    
    def transform(self, x):
        if self.target in x.columns:
            x = x.drop(columns=[self.target])
            
        if self.intercept:
            x = sm.add_constant(x)
        
        return self.classifier.predict(x)
    
    def predict(self, x):
        return self.transform(x)
    
    def summary(self):
        describe = self.classifier.summary2()
        return describe
    
    def feature_importances(self, x):
        params = {
            "n_estimators": 256,
            "max_depth": 4,
            "min_samples_split": 5,
            "learning_rate": 1e-3,
            "loss": "deviance",
            "subsample": 0.9,
        }
        feature_importances_ = GradientBoostingClassifier(**params).fit(x.drop(columns=[self.target]), x[self.target]).feature_importances_
        return pd.DataFrame(feature_importances_, index=self.feature_names_, columns=["feature_importances"])
        
    def VIF(self, x):
        if self.intercept:
            x = sm.add_constant(x)
        
        x = x.drop(columns=[self.target])
        columns = x.columns
        vif = pd.DataFrame({"VIF": [variance_inflation_factor(np.matrix(x), i) for i in range(len(columns))]}, index=columns)
        
        return vif
    
    def WALD(self):
        return self.classifier.wald_test_terms().table[["statistic", "pvalue"]].rename(columns={"pvalue": "wald_test_pvalue", "statistic": "wald_test_statistic"})
    
    def report(self):
        return self.classifier.summary2().tables[1].join([self.coef_normalization, self.WALD(), self.vif, self.feature_importances_]), self.classifier.summary2().tables[0], self.corr
    
    def summary_save(self, excel_name="逻辑回归模型拟合效果.xlsx", sheet_name="逻辑回归拟合效果"):
        writer = pd.ExcelWriter(excel_name, engine='openpyxl')
        
        coef_report, summary_report, corr_report = self.report()
        summary_report.columns = ["逻辑回归模型拟合效果"] * summary_report.shape[1]
        summary_report.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startcol=0, startrow=2)
        coef_report.reset_index().rename(columns={"index": "variable"}).to_excel(writer, sheet_name=sheet_name, index=False, header=True, startcol=0, startrow=summary_report.shape[0] + 4)
        corr_report.to_excel(writer, sheet_name=sheet_name, index=True, header=True, startcol=0, startrow=summary_report.shape[0] + coef_report.shape[0] + 7)
        
        writer.save()
        writer.close()
        
        if os.path.exists(excel_name):
            workbook = load_workbook(excel_name)
            worksheet = workbook.get_sheet_by_name(sheet_name)
            worksheet["A1"].value = "逻辑回归模型报告"
            worksheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
            worksheet.merge_cells(f"A1:L1")
            
            workbook.save(excel_name)
            workbook.close()
        
        render_excel(excel_name, sheet_name=sheet_name, max_column_width=25, merge_rows=np.cumsum([1, len(summary_report), 2, len(coef_report) + 1, 2, len(corr_report) + 1]).tolist())
        

class ITLubberLogisticRegression(LogisticRegression):
    """
    Extended Logistic Regression.
    Extends [sklearn.linear_model.LogisticRegression](https://scikit-learn.org/stable/modules/generated/sklearn.linear_model.LogisticRegression.html).
    This class provides the following extra statistics, calculated on `.fit()` and accessible via `.summary()`:
    - `cov_matrix_`: covariance matrix for the estimated parameters.
    - `std_err_intercept_`: estimated uncertainty for the intercept
    - `std_err_coef_`: estimated uncertainty for the coefficients
    - `z_intercept_`: estimated z-statistic for the intercept
    - `z_coef_`: estimated z-statistic for the coefficients
    - `p_value_intercept_`: estimated p-value for the intercept
    - `p_value_coef_`: estimated p-value for the coefficients
    Example:
    ```python
    pipeline = Pipeline([
        ('bucketer', EqualFrequencyBucketer(n_bins=10)),
        ('clf', LogisticRegression(calculate_stats=True))
    ])
    pipeline.fit(X, y)
    pipeline.named_steps['clf'].summary()
    ```
    An example output of `.summary()`:
    
    Index     | Coef.     | Std.Err  |   z       | Pz
    --------- | ----------| ---------| ----------| ------------
    const     | -0.537571 | 0.096108 | -5.593394 | 2.226735e-08
    EDUCATION | 0.010091  | 0.044874 | 0.224876  | 8.220757e-01
    """

    def __init__(self, target="target", penalty="l2", calculate_stats=True, dual=False, tol=0.0001, C=1.0, fit_intercept=True, intercept_scaling=1, class_weight=None, random_state=None, solver="lbfgs", max_iter=100, multi_class="auto", verbose=0, warm_start=False, n_jobs=None, l1_ratio=None,):
        """
        Extends [sklearn.linear_model.LogisticRegression.fit()](https://scikit-learn.org/stable/modules/generated/sklearn.linear_model.LogisticRegression.html).
        Args:
            calculate_stats (bool): If true, calculate statistics like standard error during fit, accessible with .summary()
        """
        super().__init__(penalty=penalty, dual=dual, tol=tol, C=C, fit_intercept=fit_intercept, intercept_scaling=intercept_scaling, class_weight=class_weight, random_state=random_state, solver=solver, max_iter=max_iter, multi_class=multi_class, verbose=verbose, warm_start=warm_start, n_jobs=n_jobs, l1_ratio=l1_ratio,)
        self.target = target
        self.calculate_stats = calculate_stats

    def fit(self, x, sample_weight=None, **kwargs):
        y = x[self.target]
        x = x.drop(columns=[self.target])
        
        if not self.calculate_stats:
            return super().fit(x, y, sample_weight=sample_weight, **kwargs)

        x = self.convert_sparse_matrix(x)
        if isinstance(x, pd.DataFrame):
            self.names_ = ["const"] + [f for f in x.columns]
        else:
            self.names_ = ["const"] + [f"x{i}" for i in range(x.shape[1])]

        lr = super().fit(x, y, sample_weight=sample_weight, **kwargs)

        predProbs = self.predict_proba(x)

        # Design matrix -- add column of 1's at the beginning of your x matrix
        if lr.fit_intercept:
            x_design = np.hstack([np.ones((x.shape[0], 1)), x])
        else:
            x_design = x

        p = np.product(predProbs, axis=1)
        self.cov_matrix_ = np.linalg.inv((x_design * p[..., np.newaxis]).T @ x_design)
        std_err = np.sqrt(np.diag(self.cov_matrix_)).reshape(1, -1)

        # In case fit_intercept is set to True, then in the std_error array
        # Index 0 corresponds to the intercept, from index 1 onwards it relates to the coefficients
        # If fit intercept is False, then all the values are related to the coefficients
        if lr.fit_intercept:

            self.std_err_intercept_ = std_err[:, 0]
            self.std_err_coef_ = std_err[:, 1:][0]

            self.z_intercept_ = self.intercept_ / self.std_err_intercept_

            # Get p-values under the gaussian assumption
            self.p_val_intercept_ = scipy.stats.norm.sf(abs(self.z_intercept_)) * 2

        else:
            self.std_err_intercept_ = np.array([np.nan])
            self.std_err_coef_ = std_err[0]

            self.z_intercept_ = np.array([np.nan])

            # Get p-values under the gaussian assumption
            self.p_val_intercept_ = np.array([np.nan])

        self.z_coef_ = self.coef_ / self.std_err_coef_
        self.p_val_coef_ = scipy.stats.norm.sf(abs(self.z_coef_)) * 2

        return self

    def summary(self):
        """
        Puts the summary statistics of the fit() function into a pandas DataFrame.
        Returns:
            data (pandas DataFrame): The statistics dataframe, indexed by the column name
        """
        check_is_fitted(self)

        if not hasattr(self, "std_err_coef_"):
            msg = "Summary statistics were not calculated on .fit(). Options to fix:\n"
            msg += "\t- Re-fit using .fit(X, y, calculate_stats=True)\n"
            msg += "\t- Re-inititialize using LogisticRegression(calculate_stats=True)"
            raise AssertionError(msg)

        data = {
            "Coef.": (self.intercept_.tolist() + self.coef_.tolist()[0]),
            "Std.Err": (self.std_err_intercept_.tolist() + self.std_err_coef_.tolist()),
            "z": (self.z_intercept_.tolist() + self.z_coef_.tolist()[0]),
            "P>|z|": (self.p_val_intercept_.tolist() + self.p_val_coef_.tolist()[0]),
        }
        
        stats = pd.DataFrame(data, index=self.names_)
        stats["[ 0.025"] = stats["Coef."] - 1.96 * stats["Std.Err"]
        stats["0.975 ]"] = stats["Coef."] + 1.96 * stats["Std.Err"]
        
        return stats
    
    @staticmethod
    def convert_sparse_matrix(x):
        """
        Converts a sparse matrix to a numpy array.
        This can prevent problems arising from, e.g. OneHotEncoder.
        Args:
            x: numpy array, sparse matrix
        Returns:
            numpy array of x
        """
        if scipy.sparse.issparse(x):
            return x.toarray()
        else:
            return x

    def plot_weights(self, format=None, scale=None, width=None, height=None):
        """
        Generates a weight plot(plotly chart) from `stats`
        Example:
        ```
        pipeline = Pipeline([
            ('bucketer', EqualFrequencyBucketer(n_bins=10)),
            ('clf', LogisticRegression(calculate_stats=True))
        ])
        pipeline.fit(X, y)
        stats = pipeline.named_steps['clf'].plot_weights()
        ```
        Args:
            stats: The statistics to display
            format: The format of the image, such as 'png'. The default None returns a plotly image.
            scale: If format is specified, the scale of the image
            width: If format is specified, the width of the image
            height: If format is specified, the image of the image
        """
        stats = self.summary()
        
        fig = go.Figure()

        fig.add_trace(
            go.Scatter(
                x=stats['Coef.'],
                y=stats['Coef.'].index,
                line=dict(color='#2639E9', width=2),
                mode='markers',

                error_x=dict(
                    type='data',
                    symmetric=False,
                    array=stats['0.975 ]'] - stats['Coef.'],
                    arrayminus=stats['Coef.'] - stats['[ 0.025'],
                    color='#2639E9')
            )
        )

        fig.add_shape(type="line",
                    x0=0, y0=0, x1=0, y1=len(stats),
                    line=dict(color="#a29bfe", width=3, dash='dash')
                    )

        fig.update_layout(
            title='Regression Meta Analysis - Weight Plot',
            xaxis_title='Weight Estimates',
            yaxis_title='Variable',
            xaxis_showgrid=False,
            yaxis_showgrid=False
        )
        
        fig.update_layout(template="simple_white")

        if format is not None:
            img_bytes = fig.to_image(format=format, scale=scale, width=width, height=height)
            fig = Image(img_bytes)
            
        return fig
    
    
class ScoreCard(TransformerMixin, BaseEstimator, ClassifierMixin):
    
    def __init__(self, target="target", pdo=60, rate=2, base_odds=35, base_score=750, combiner={}, transer=None, penalty='l2', C=1.0, fit_intercept=True, class_weight="balanced", random_state=None, 
                 solver="lbfgs", max_iter=100, verbose=0, warm_start=False, n_jobs=None, l1_ratio=None, ):
        self.target = target
        self.combiner = combiner
        self.transer = transer
        self.base_score = base_score
        self.base_odds = base_odds
        self.rate = rate
        self.pdo = pdo
        self.penalty = penalty
        self.C = C
        self.fit_intercept = fit_intercept
        self.class_weight = class_weight
        self.random_state = random_state
        self.solver = solver
        self.max_iter = max_iter
        self.verbose = verbose
        self.warm_start = warm_start
        self.n_jobs = n_jobs
        self.l1_ratio = l1_ratio
        self.card = toad.ScoreCard(combiner=self.combiner, transer=self.transer, pdo=self.pdo, rate=self.rate, base_odds=self.base_odds, base_score=self.base_score, penalty=self.penalty, C=self.C, 
                                   fit_intercept=self.fit_intercept, class_weight=self.class_weight, random_state=self.random_state, solver=self.solver, max_iter=self.max_iter, verbose=self.verbose, 
                                   warm_start=self.warm_start, n_jobs=self.n_jobs, l1_ratio=self.l1_ratio, )
        
    def fit(self, x, y=None):
        self.classes_ = x[self.target].unique()
        self.card.fit(x.drop(columns=[self.target]), x[self.target])
        return self
    
    def transform(self, x):
        return self.card.predict(x)
    
    def predict(self, x):
        return self.card.predict(x)
    
    def export(self):
        return self.card.export()
    
    def KS_bucket(self, y_pred, y_true, bucket=10, method="quantile"):
        return toad.metrics.KS_bucket(y_pred, y_true, bucket=bucket, method=method)
    
    def KS(self, y_pred, y_true):
        return toad.metrics.KS(y_pred, y_true)
    
    def AUC(self, y_pred, y_true):
        return toad.metrics.AUC(y_pred, y_true)
    
    def perf_eva(self, y_pred, y_true, title="", plot_type=["ks", "roc"]):
        return sc.perf_eva(y_true, y_pred, title=title, plot_type=plot_type)
    
    def PSI(self, y_pred_train, y_pred_oot):
        return toad.metrics.PSI(y_pred_train, y_pred_oot)
    
    def perf_psi(self, y_pred_train, y_pred_oot, y_true_train, y_true_oot, keys=["train", "test"], x_limits=None, x_tick_break=50, show_plot=True, return_distr_dat=False):
        return sc.perf_psi(
            score = {keys[0]: y_pred_train, keys[1]: y_pred_oot},
            label = {keys[0]: y_true_train, keys[1]: y_true_oot},
            x_limits = x_limits,
            x_tick_break = x_tick_break,
            show_plot = show_plot,
            return_distr_dat = return_distr_dat,
        )
    
    def score_hist(self, score, y_true, figsize=(15, 10), bins=20, alpha=0.6):
        mask = y_true == 0
        fig = plt.figure(figsize=figsize)
        plt.hist(score[mask], label="好样本", color="#2639E9", alpha=alpha, bins=bins)
        plt.hist(score[~mask], label="坏样本", color="#F76E6C", alpha=alpha, bins=bins)
        plt.xlabel("score")
        plt.legend()
        plt.show()
        
        return fig


if __name__ == "__main__":
    target = "creditability"
    data = sc.germancredit()
    data[target] = data[target].map({"good": 0, "bad": 1})
    
    train, test = train_test_split(data, test_size=0.3, shuffle=True, stratify=data[target])
    
    # selection = FeatureSelection(target=target, engine="toad", return_drop=True, corr=0.9, iv=0.01)
    # train = selection.fit_transform(train)
    
    # combiner = Combiner(min_samples=0.2, empty_separate=True, target=target)
    # combiner.fit(train)
    # train = combiner.transform(train)
    
    # transformer = WOETransformer(target=target)
    # train = transformer.fit_transform(train)
    
    # stepwise = StepwiseSelection(target=target)
    # train = stepwise.fit_transform(train)
    
    feature_pipeline = Pipeline([
        ("preprocessing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("combiner", Combiner(target=target, min_samples=0.2)),
        ("transformer", WOETransformer(target=target)),
        ("processing_select", FeatureSelection(target=target, engine="scorecardpy")),
        ("stepwise", StepwiseSelection(target=target, target_rm=False)),
        # ("logistic", LogisticClassifier(target=target)),
        # ("logistic", ITLubberLogisticRegression(target=target)),
    ])
    
    # feature_pipeline.fit(train)
    # y_pred_train = feature_pipeline.predict(train.drop(columns=target))
    # y_pred_test = feature_pipeline.predict(test.drop(columns=target))

    # params_grid = {
    #     # "logistic__C": [i / 1. for i in range(1, 10, 2)],
    #     # "logistic__penalty": ["l2"],
    #     # "logistic__class_weight": [None, "balanced"], # + [{1: i / 10.0, 0: 1 - i / 10.0} for i in range(1, 10)],
    #     # "logistic__max_iter": [100],
    #     # "logistic__solver": ["sag"] # ["liblinear", "sag", "lbfgs", "newton-cg"],
    #     "logistic__intercept": [True, False],
    # }
    
    # clf = GridSearchCV(feature_pipeline, params_grid, cv=5, scoring='roc_auc', verbose=-1, n_jobs=2, return_train_score=True)
    # clf.fit(train, train[target])
    
    # y_pred_train = clf.best_estimator_.predict(train)
    # y_pred_test = clf.best_estimator_.predict(test)
    
    # print(clf.best_params_)
    
    # statmodels methods
    # feature_pipeline.named_steps['logistic'].summary_save()
    
    # print("train: ", toad.metrics.KS(y_pred_train, train[target]), toad.metrics.AUC(y_pred_train, train[target]))
    # print("test: ", toad.metrics.KS(y_pred_test, test[target]), toad.metrics.AUC(y_pred_test, test[target]))
    
    woe_train = feature_pipeline.fit_transform(train)
    woe_test = feature_pipeline.transform(test)
    
    # lr = LogisticClassifier(target=target)
    # lr.fit(woe_train)
    # lr.summary_save()

    cols = list(filter(lambda x: x != target, feature_pipeline.named_steps['preprocessing_select'].select_columns))
    combiner = feature_pipeline.named_steps['combiner'].combiner
    transformer = feature_pipeline.named_steps['transformer'].transformer
    
    # feature_describe = pd.read_excel("变量字典及字段解释.xlsx", sheet_name="数据字段表", header=0, engine="openpyxl", usecols=[0, 1])
    # feature_describe = feature_describe.drop_duplicates(subset=["变量名称"], keep="last")
    # feature_dict = dict(zip(feature_describe["变量名称"], feature_describe["含义"]))

    # # 保存结果至 EXCEL 文件
    # output_excel_name = "指标有效性验证.xlsx"
    # output_sheet_name = "指标有效性"
    # tables = {}
    # merge_row_number = []

    # for feature in cols:
    #     table = feature_bin_stats(train, feature, feature_dict=feature_dict, rules={}, target=target)
    #     df_psi = cal_psi(train[[feature, target]], test[[feature, target]], feature, combiner=combiner)
        
    #     table = table.merge(df_psi, on="分箱", how="left")
        
    #     feature_bin = combiner.export()[feature]
    #     feature_bin_dict = format_bins(np.array(feature_bin))
    #     table["分箱"] = table["分箱"].map(feature_bin_dict)
        
    #     merge_row_number.append(len(table))
    #     tables[feature] = table

    # merge_row_number = np.cumsum(merge_row_number).tolist()
    # feature_table = pd.concat(tables, ignore_index=True).round(6)
    # feature_table["分档WOE值"] = feature_table["分档WOE值"].fillna(np.inf)
    # feature_table.to_excel(output_excel_name, sheet_name=output_sheet_name, index=False, header=True, startcol=0, startrow=0)

    # render_excel(output_excel_name, sheet_name=output_sheet_name, conditional_columns=["J", "N"], freeze="D2", merge_rows=merge_row_number, percent_columns=[5, 7, 9, 10])
        
    score_card = ScoreCard(target=target, combiner=combiner, transer=transformer, )
    score_card.fit(woe_train)
    
    
    data["score"] = score_card.transform(data)
    
    print(score_card.KS_bucket(data["score"], data[target]))
    pt = score_card.perf_eva(data["score"], data[target], title="train")
    
    sc = score_card.score_hist(data["score"], data[target])
    
    print(score_card.KS(data["score"], data[target]), score_card.AUC(data["score"], data[target]))
    